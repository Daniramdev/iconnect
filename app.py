from flask import render_template,request,redirect,session,url_for,flash,jsonify, make_response, flash,send_file
from flask import Flask
import os
import pandas as pd
from passlib.hash import sha256_crypt
import mysql.connector
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import datetime
import openpyxl
from flask import Response
from io import BytesIO, StringIO
import uuid  # Add this line
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
import numpy as np 
from decimal import Decimal
import csv
import shutil
import logging




app = Flask(__name__)
app.secret_key = "your_key bebas"

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Fungsi untuk mengenkripsi password
def encrypt_password(password):
    return sha256_crypt.encrypt(password)


# Fungsi untuk memverifikasi password
def verify_password(password, hashed_password):
    return sha256_crypt.verify(password, hashed_password)


# Fungsi untuk melakukan koneksi ke database MySQL
def get_db_connection():
    mydb = mysql.connector.connect(
        host="localhost", user="root", password="", database="db_project1"
    )
    return mydb


@app.route("/")
def index():
    return render_template("registrasi/login.html")


# ----------------LOGIN----------------------
# Halaman register
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"]
        email = request.form["email"]
        password = request.form["password"]
        confirm_password = request.form["confirm_password"]

        # Check if passwords match
        if password != confirm_password:
            error = "Passwords do not match!"
            return render_template("registrasi/register.html", error=error)

        # Check if password length is less than 2 characters
        if len(password) < 3:
            error = "Password must be at least 2 characters long!"
            return render_template("registrasi/register.html", error=error)

        hashed_password = encrypt_password(password)

        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Cek apakah username atau email sudah terdaftar
        cursor.execute(
            "SELECT * FROM users WHERE username = %s OR email = %s", (username, email)
        )
        user = cursor.fetchone()
        if user:
            error = "Username or email already exists!"
            return render_template("registrasi/register.html", error=error)
        else:
            # Tambahkan user baru ke database
            cursor.execute(
                "INSERT INTO users (username, email, password) VALUES (%s, %s, %s)",
                (username, email, hashed_password),
            )
            mydb.commit()
            cursor.close()
            mydb.close()
            # Pass a success message to the login page
            success_message = "Registration successful! You can now log in."
            return redirect(url_for("login", success_message=success_message))

    return render_template("registrasi/register.html")




# Halaman login
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Ambil data user dari database
        cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()

        if user:
            hashed_password = user[3]
            # Verifikasi password
            if verify_password(password, hashed_password):
                # Set session untuk user yang berhasil login
                session["email"] = email
                session["username"] = user[1]
                session["role"] = user[4]
                cursor.close()
                mydb.close()
                session["login"] = True
                flash("Login successful!", "success")
                return redirect("dashboard")
            else:
                error = "Invalid password!"
                return render_template("/registrasi/login.html", error=error)
        else:
            error = "Email not found!"
            return render_template("/registrasi/login.html", error=error)

    return render_template("/registrasi/login.html")


@app.route('/search_menu', methods=['GET'])
def search_menu():
    search_query = request.args.get('query', '').lower()

    # Data menu dapat diambil dari sumber data Anda
    # Di sini saya memberikan contoh menu statis
    menu_data = [
        {"url": "/admin_dashboard", "text": "Home"},
        {"url": "/assets", "text": "Dashboard Assets"},
        {"url": "/dashboard", "text": "Dashboard Retail"},
        # ... (tambahkan data submenu lainnya) ...
    ]

    filtered_menu = [item for item in menu_data if search_query in item['text'].lower()]

    return jsonify(filtered_menu)


@app.route("/dashboard")
def dashboard():
    # Periksa apakah ada pesan logout di sesi
    logout_message = session.pop("logout_message", None)

    # Jika ada pesan logout, tampilkan peringatan
    if logout_message:
        flash(logout_message, "warning")

    # Check if the "role" key exists in the session
    if "role" in session:
        role = session["role"]

        # Check the role and redirect accordingly
        if role == "admin":
            return render_admin_dashboard()
        elif role == "user":
            return render_user_dashboard()
        else:
            # Handle unexpected roles
            return "Unknown role"
    else:
        # Redirect to login if the "role" key is not present in the session
        return redirect("/login")


# Function to calculate sums
def calculate_sums(data):
    hcbandung_sum = sum(row["bandung"] for row in data)
    hccirebon_sum = sum(row["cirebon"] for row in data)
    hctasik_sum = sum(row["tasik"] for row in data)
    
    # Sum all values across all rows for each city
    totalhc = hcbandung_sum + hccirebon_sum + hctasik_sum
    
    return hcbandung_sum, hccirebon_sum, hctasik_sum, totalhc

@app.route("/admin_dashboard")
def render_admin_dashboard():
    # Check if the user is logged in
    if 'username' in session:
        try:
            # Get the database connection
            mydb = get_db_connection()
            cursor = mydb.cursor(dictionary=True)

            # Get user information from the session
            user_name = session.get('username')
            user_role = session.get('role')

            # Fetch data for the admin dashboard
            # Your SQL query
            sql_query = "SELECT tahun, bandung, cirebon, tasik FROM tb_jbbmilestone"
            cursor.execute(sql_query)

            # Fetching data
            jbb = cursor.fetchall()

            # Calculating sums for JBB Milestone Data
            hcbandung_sum, hccirebon_sum, hctasik_sum, totalhc = calculate_sums(jbb)

            # Fetch additional data from jbbaging
            cursor.execute("SELECT kabagingbdg, kpibdg, agingbdg, kabagingcrb, kpicrb, agingcrb, kabagingtsk, kpitsk, agingtsk FROM jbbaging")
            jbbaging = cursor.fetchall()
            jbbagingbdg_sum = sum(row["agingbdg"] for row in jbbaging)
            jbbagingcrb_sum = sum(row["agingcrb"] for row in jbbaging)
            jbbagingtsk_sum = sum(row["agingtsk"] for row in jbbaging)

            # Fetch additional data from tb_akv
            cursor.execute("SELECT monthly_bulan,monthly_spa,mothlyhc_bulan,monthly_hc,akumulasihcvsspa_bulan,akumulasihcvsspa_hc,akumulasihcvsspa_target,akumulasihcvsspa_capaian,monthlyhc_bulan,monthlyhc_bdg,monthlyhc_crb,monthlyhc_tsk,akumulasihckp_bulan,akumulasihckp_bdg,akumulasihckp_crb,akumulasihckp_tsk,agingkp_bulan,agingkp_bdg,agingkp_crb,agingkp_tsk,agingkpbai_bulan,agingkpbai_averageaging FROM tb_akv")
            tb_akv_data = cursor.fetchall()
            totalavg = sum(row["agingkpbai_averageaging"] for row in tb_akv_data)

            # Close the database connection
            cursor.close()
            mydb.close()

            # Get the current date
            current_date = datetime.datetime.now()
            formatted_date = current_date.strftime("%d %B %Y")

            # Check for logout message
            logout_message = session.pop("logout_message", None)

            # If there is a logout message, display the warning
            if logout_message:
                flash(logout_message, "warning")

            # Render the admin dashboard template
            return render_template("dashboard/admin/dashboard_admin.html", user_name=user_name,
                                hcbandung_sum=hcbandung_sum, hccirebon_sum=hccirebon_sum,
                                hctasik_sum=hctasik_sum, totalhc=totalhc,
                                jbbagingbdg_sum=jbbagingbdg_sum, jbbagingcrb_sum=jbbagingcrb_sum,
                                jbbagingtsk_sum=jbbagingtsk_sum,
                                totalavg=totalavg,
                                user_role=user_role, formatted_date=formatted_date,
                                jbb=jbb, jbbaging=jbbaging)
        except Exception as e:
            # Log the exception or handle it in a way that makes sense for your application
            flash('An error occurred. Please try again later.', 'error')
            return redirect(url_for('login'))
    else:
        # Check for logout message
        logout_message = session.pop("logout_message", None)

        # If there is a logout message, display the warning
        if logout_message:
            flash(logout_message, "warning")

        # Display an error message and redirect to the login page
        flash('You need to log in first.', 'error')
        return redirect(url_for('login'))


@app.route("/user_dashboard")
def render_user_dashboard():
    # Check if the user is logged in
    if 'username' in session:
        try:
            # Get the database connection
            mydb = get_db_connection()
            cursor = mydb.cursor(dictionary=True)

            # Get user information from the session
            user_name = session.get('username')
            user_role = session.get('role')

            # Fetch data for the admin dashboard
            # Your SQL query
            sql_query = "SELECT tahun, bandung, cirebon, tasik FROM tb_jbbmilestone"
            cursor.execute(sql_query)

            # Fetching data
            jbb = cursor.fetchall()

            # Calculating sums for JBB Milestone Data
            hcbandung_sum, hccirebon_sum, hctasik_sum, totalhc = calculate_sums(jbb)

            # Fetch additional data from jbbaging
            cursor.execute("SELECT kabagingbdg, kpibdg, agingbdg, kabagingcrb, kpicrb, agingcrb, kabagingtsk, kpitsk, agingtsk FROM jbbaging")
            jbbaging = cursor.fetchall()
            jbbagingbdg_sum = sum(row["agingbdg"] for row in jbbaging)
            jbbagingcrb_sum = sum(row["agingcrb"] for row in jbbaging)
            jbbagingtsk_sum = sum(row["agingtsk"] for row in jbbaging)

            # Fetch additional data from tb_akv
            cursor.execute("SELECT monthly_bulan,monthly_spa,mothlyhc_bulan,monthly_hc,akumulasihcvsspa_bulan,akumulasihcvsspa_hc,akumulasihcvsspa_target,akumulasihcvsspa_capaian,monthlyhc_bulan,monthlyhc_bdg,monthlyhc_crb,monthlyhc_tsk,akumulasihckp_bulan,akumulasihckp_bdg,akumulasihckp_crb,akumulasihckp_tsk,agingkp_bulan,agingkp_bdg,agingkp_crb,agingkp_tsk,agingkpbai_bulan,agingkpbai_averageaging FROM tb_akv")
            tb_akv_data = cursor.fetchall()
            totalavg = sum(row["agingkpbai_averageaging"] for row in tb_akv_data)

            # Close the database connection
            cursor.close()
            mydb.close()

            # Get the current date
            current_date = datetime.datetime.now()
            formatted_date = current_date.strftime("%d %B %Y")

            # Check for logout message
            logout_message = session.pop("logout_message", None)

            # If there is a logout message, display the warning
            if logout_message:
                flash(logout_message, "warning")

            # Render the admin dashboard template
            return render_template("dashboard/user/dashboard_users.html", user_name=user_name,
                                hcbandung_sum=hcbandung_sum, hccirebon_sum=hccirebon_sum,
                                hctasik_sum=hctasik_sum, totalhc=totalhc,
                                jbbagingbdg_sum=jbbagingbdg_sum, jbbagingcrb_sum=jbbagingcrb_sum,
                                jbbagingtsk_sum=jbbagingtsk_sum,
                                totalavg=totalavg,
                                user_role=user_role, formatted_date=formatted_date,
                                jbb=jbb, jbbaging=jbbaging)
        except Exception as e:
            # Log the exception or handle it in a way that makes sense for your application
            flash('An error occurred. Please try again later.', 'error')
            return redirect(url_for('login'))
    else:
        # Check for logout message
        logout_message = session.pop("logout_message", None)

        # If there is a logout message, display the warning
        if logout_message:
            flash(logout_message, "warning")

        # Display an error message and redirect to the login page
        flash('You need to log in first.', 'error')
        return redirect(url_for('login'))
    


@app.route("/sales")
def sales():
    try:
        # Establish a database connection
        mydb = get_db_connection()

        if mydb is None:
            # Handle the case where the connection couldn't be established
            return "Unable to connect to the database", 500

        cursor = mydb.cursor(dictionary=True)

        # Query for total revenue
    # Query for total revenue
        cursor.execute("SELECT * FROM tb_rawdata_retail")
        assets = cursor.fetchall()
        total_revenue = len(assets)
        print("Jumlah Baris:", total_revenue)

        # Query for total values for hc, rev, and accumm_rev
        cursor.execute("SELECT montly_rev_bulan, montly_rev_hc, montly_rev, accumm_rev FROM tb_retail")
        total = cursor.fetchall()
        total_hc = sum(row["montly_rev_hc"] for row in total)

        total_rev = sum(row["montly_rev"] for row in total)
        formatted_total_rev = f"{total_rev:,.0f}"
        total_accumm_rev = sum(row["accumm_rev"] for row in total)
        formatted_total_accumm_rev = f"{total_accumm_rev:,.0f}"


        cursor.close()
        mydb.close()

        current_date = datetime.datetime.now()
        formatted_date = current_date.strftime("%d %B %Y")

        return render_template(
            "dashboard/admin/dashboard_sales.html", 
            total_revenue=total_revenue,
            formatted_date=formatted_date,
            total_hc=total_hc,
            total_rev=formatted_total_rev,
            total_accumm_rev=formatted_total_accumm_rev,
        )

    except Exception as e:
        return str(e), 500



@app.route("/salesuser")
def salesuser():
    try:
        # Establish a database connection
        mydb = get_db_connection()

        if mydb is None:
            # Handle the case where the connection couldn't be established
            return "Unable to connect to the database", 500

        cursor = mydb.cursor(dictionary=True)

        # Query for total revenue
    # Query for total revenue
        cursor.execute("SELECT * FROM tb_rawdata_retail")
        assets = cursor.fetchall()
        total_revenue = len(assets)
        print("Jumlah Baris:", total_revenue)

        # Query for total values for hc, rev, and accumm_rev
        cursor.execute("SELECT montly_rev_bulan, montly_rev_hc, montly_rev, accumm_rev FROM tb_retail")
        total = cursor.fetchall()
        total_hc = sum(row["montly_rev_hc"] for row in total)

        total_rev = sum(row["montly_rev"] for row in total)
        formatted_total_rev = f"{total_rev:,.0f}"
        total_accumm_rev = sum(row["accumm_rev"] for row in total)
        formatted_total_accumm_rev = f"{total_accumm_rev:,.0f}"


        cursor.close()
        mydb.close()

        current_date = datetime.datetime.now()
        formatted_date = current_date.strftime("%d %B %Y")

        return render_template(
            "dashboard/admin/dashboard_sales.html", 
            total_revenue=total_revenue,
            formatted_date=formatted_date,
            total_hc=total_hc,
            total_rev=formatted_total_rev,
            total_accumm_rev=formatted_total_accumm_rev,
        )

    except Exception as e:
        return str(e), 500

# ------------------ENDLOGIN----------------------


# ----------------------Crud----------------------
@app.route("/userlist")
def userlist():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT id, username, email, role FROM users")
    user = cursor.fetchall()
    cursor.close()

    return render_template("/managementuser/user_list.html", user=user)

@app.route('/search', methods=['GET'])
def search():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    
    query = request.args.get('q', '').lower()

    cursor.execute("SELECT id, username, email, role FROM users")
    user = cursor.fetchall()
    
    search_results = [usr for usr in user if
                    query in usr['username'].lower() or
                    query in usr['email'].lower() or
                    query in usr['role'].lower()]

    return render_template('managementuser/user_list.html', user=search_results)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    bcrypt = Bcrypt()

    if request.method == 'GET':
        cursor.execute("SELECT id, username, email, role FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if user:
            return render_template("managementuser/user_list.html", user=user)
        else:
            flash('User tidak ditemukan', 'error')
            return redirect(url_for('userlist'))

    elif request.method == 'POST':
        # Handle the form update
        new_username = request.form['username']
        new_email = request.form['email']
        new_role = request.form['role']
        new_password = request.form['password']

        # Check if a new password is provided before updating it
        if new_password:
            hashed_password = sha256_crypt.using(rounds=1000).hash(new_password)
            cursor.execute("UPDATE users SET username = %s, email = %s, role = %s, password = %s WHERE id = %s",
                           (new_username, new_email, new_role, hashed_password, user_id))
        else:
            # If no new password is provided, update other fields without changing the password
            cursor.execute("UPDATE users SET username = %s, email = %s, role = %s WHERE id = %s",
                           (new_username, new_email, new_role, user_id))

        mydb.commit()

        flash('User berhasil diupdate', 'success')

        cursor.close()
        mydb.close()

        return redirect(url_for('userlist'))

    return "Metode HTTP tidak valid", 405


@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    mydb = get_db_connection()
    cursor = mydb.cursor()

    try:
        # Hapus pengguna berdasarkan ID
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        mydb.commit()
        return redirect('/userlist')

    except Exception as e:
        # Tangani kesalahan, bisa dicetak atau disimpan ke log
        print(f"Error deleting user: {e}")
        mydb.rollback()

    finally:
        cursor.close()
        mydb.close()

        

def get_sales_data(page, per_page):
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    # Assuming 'tb_rawdata_retail' is the table name
    cursor.execute("SELECT * FROM tb_rawdata_retail")
    total_sales = cursor.fetchall()

    start = (page - 1) * per_page
    end = start + per_page
    sales = total_sales[start:end]

    cursor.close()

    return sales


@app.route("/tb_retail", methods=['GET'])
def tb_retail():
    # Get the page number from the request, default to 1 if not provided
    page = request.args.get('page', 1, type=int)
    per_page = 1000  # Adjust this number as needed

    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    offset = (page - 1) * per_page
    cursor.execute(f"SELECT * FROM data_retail LIMIT {per_page} OFFSET {offset}")
    retail = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_retail/tb_retailbig.html", retail=retail)


# Export route
@app.route("/tb_retail/export_excel", methods=['GET'])
def export_excel_tb_retail():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT * FROM data_retail")
    retail = cursor.fetchall()
    cursor.close()

    # Convert the data to a DataFrame
    df = pd.DataFrame(retail)

    # Save the DataFrame to an Excel file
    export_path = "tb_retail_export.xlsx"
    df.to_excel(export_path, index=False)

    # Send the file as a response
    return send_file(export_path, as_attachment=True)






# @app.route("/import_tb_sales", methods=['POST'])
# def import_tb_sales():
#     file = request.files['file']
#     if file:
#         # Read the Excel file into a Pandas DataFrame
#         df = pd.read_excel(file)

#         # Ensure that the DataFrame columns match the expected columns
#         expected_columns = []

#         if not all(column in df.columns for column in expected_columns):
#             return "Error: Columns in the uploaded file do not match the expected columns."

#         # Handle NaN values in the DataFrame by replacing them with None
#         df = df.applymap(lambda x: None if pd.isna(x) else x)

#         # You can now process the DataFrame as needed, for example, insert into the database
#         mydb = get_db_connection()
#         cursor = mydb.cursor()

#         # Assuming 'tb_sales' is the table name
#         # Using executemany to insert multiple rows at once
#         columns_str = ', '.join(expected_columns)
#         placeholders = ', '.join(['%s' for _ in expected_columns])
#         query = f"INSERT INTO data_retail ({columns_str}) VALUES ({placeholders})"
        
#         try:
#             cursor.executemany(query, df.values.tolist())
#             mydb.commit()
#             return "File uploaded and data imported successfully."
#         except Exception as e:
#             return f"Error: {e}"

#         finally:
#             cursor.close()

#     return "No file provided."



@app.route("/assets")
def assets():
    # Establish a database connection
    mydb = get_db_connection()

    if mydb is None:
        # Handle the case where the connection couldn't be established
        return "Unable to connect to the database", 500

    cursor = mydb.cursor(dictionary=True)

    try:
        # Perform database queries for oltbdg
        cursor.execute(
            "SELECT totaloltbdg, totalportbdg, idleportbdg, totaloltcrb,totalportcrb, idleportcrb, totalolttsk,totalporttsk, idleporttsk FROM tb_olt"
        )
        olt = cursor.fetchall()

        # Calculate the total sum for totaloltbdg, totaloltcrb, and totalolttsk
        totalSum = sum(
            item["totaloltbdg"] + item["totaloltcrb"] + item["totalolttsk"]
            for item in olt)

        # Perform a separate query to get the sums directly
        cursor.execute("SELECT SUM(totaloltbdg) AS totaloltbdg_sum, SUM(totaloltcrb) AS totaloltcrb_sum,SUM(totalolttsk) AS totalolttsk_sum, SUM(totalportbdg) AS totalportbdg_sum,SUM(totalportcrb) AS totalportcrb_sum,SUM(totalporttsk) AS totalporttsk_sum,SUM(idleportbdg) AS idleportbdg_sum,SUM(idleportcrb) AS idleportcrb_sum,SUM(idleporttsk) AS idleporttsk_sum  FROM tb_olt")
        total_sum_result = cursor.fetchone()

        totaloltbdg_sum = total_sum_result['totaloltbdg_sum'] if total_sum_result else 0
        totaloltcrb_sum = total_sum_result['totaloltcrb_sum'] if total_sum_result else 0
        totalolttsk_sum = total_sum_result['totalolttsk_sum'] if total_sum_result else 0
        totalportbdg_sum = total_sum_result['totalportbdg_sum'] if total_sum_result else 0
        totalportcrb_sum = total_sum_result['totalportcrb_sum'] if total_sum_result else 0
        totalporttsk_sum = total_sum_result['totalporttsk_sum'] if total_sum_result else 0
        idleportbdg_sum = total_sum_result['idleportbdg_sum'] if total_sum_result else 0
        idleportcrb_sum = total_sum_result['idleportcrb_sum'] if total_sum_result else 0
        idleporttsk_sum = total_sum_result['idleporttsk_sum'] if total_sum_result else 0



        cursor.execute("SELECT hpbdg, hcbdg, turbdg, hpcrb, hccrb, turcrb, hptsk, hctsk, turtsk FROM tb_tur")
        # Calculate the total sum for totaloltbdg, totaloltcrb, and totalolttsk
        rjbb = cursor.fetchall()

        totalhp = sum(
            item["hpbdg"] + item["hpcrb"] + item["hptsk"]
            for item in rjbb
        )

        totalhc = sum(
            item["hcbdg"] + item["hccrb"] + item["hctsk"]
            for item in rjbb
        )
        totaltur = sum(
            item["turbdg"] + item["turcrb"] + item["turtsk"]
            for item in rjbb
        )
        formatted_totaltur = "{:02.0f}".format(totaltur)

        print(formatted_totaltur)

        cursor.execute("SELECT bulan_hcvstarget, akumulasihc, target, pencapaian FROM tb_retail")
        homepass = cursor.fetchall()

        cursor.close()
        mydb.close()

        return render_template(
            "dashboard/admin/dashboard_assets.html", olt=olt,totalhp=totalhp,totalhc=totalhc,formatted_totaltur=formatted_totaltur, totalSum=totalSum, homepass=homepass, totaloltbdg_sum=totaloltbdg_sum,totaloltcrb_sum=totaloltcrb_sum,totalolttsk_sum=totalolttsk_sum,totalportbdg_sum=totalportbdg_sum,totalportcrb_sum=totalportcrb_sum,totalporttsk_sum=totalporttsk_sum,idleportbdg_sum=idleportbdg_sum,idleportcrb_sum=idleportcrb_sum,idleporttsk_sum=idleporttsk_sum,
        )
    except Exception as e:
        return str(e), 500



    
@app.route("/assetsuser")
def assetsuser():
    # Establish a database connection
    mydb = get_db_connection()

    if mydb is None:
        # Handle the case where the connection couldn't be established
        return "Unable to connect to the database", 500

    cursor = mydb.cursor(dictionary=True)

    try:
        # Perform database queries for oltbdg
        cursor.execute(
            "SELECT totaloltbdg, totalportbdg, idleportbdg, totaloltcrb,totalportcrb, idleportcrb, totalolttsk,totalporttsk, idleporttsk FROM tb_olt"
        )
        olt = cursor.fetchall()

        # Calculate the total sum for totaloltbdg, totaloltcrb, and totalolttsk
        totalSum = sum(
            item["totaloltbdg"] + item["totaloltcrb"] + item["totalolttsk"]
            for item in olt)

        # Perform a separate query to get the sums directly
        cursor.execute("SELECT SUM(totaloltbdg) AS totaloltbdg_sum, SUM(totaloltcrb) AS totaloltcrb_sum,SUM(totalolttsk) AS totalolttsk_sum, SUM(totalportbdg) AS totalportbdg_sum,SUM(totalportcrb) AS totalportcrb_sum,SUM(totalporttsk) AS totalporttsk_sum,SUM(idleportbdg) AS idleportbdg_sum,SUM(idleportcrb) AS idleportcrb_sum,SUM(idleporttsk) AS idleporttsk_sum  FROM tb_olt")
        total_sum_result = cursor.fetchone()

        totaloltbdg_sum = total_sum_result['totaloltbdg_sum'] if total_sum_result else 0
        totaloltcrb_sum = total_sum_result['totaloltcrb_sum'] if total_sum_result else 0
        totalolttsk_sum = total_sum_result['totalolttsk_sum'] if total_sum_result else 0
        totalportbdg_sum = total_sum_result['totalportbdg_sum'] if total_sum_result else 0
        totalportcrb_sum = total_sum_result['totalportcrb_sum'] if total_sum_result else 0
        totalporttsk_sum = total_sum_result['totalporttsk_sum'] if total_sum_result else 0
        idleportbdg_sum = total_sum_result['idleportbdg_sum'] if total_sum_result else 0
        idleportcrb_sum = total_sum_result['idleportcrb_sum'] if total_sum_result else 0
        idleporttsk_sum = total_sum_result['idleporttsk_sum'] if total_sum_result else 0



        cursor.execute("SELECT hpbdg, hcbdg, turbdg, hpcrb, hccrb, turcrb, hptsk, hctsk, turtsk FROM tb_tur")
        # Calculate the total sum for totaloltbdg, totaloltcrb, and totalolttsk
        rjbb = cursor.fetchall()

        totalhp = sum(
            item["hpbdg"] + item["hpcrb"] + item["hptsk"]
            for item in rjbb
        )

        totalhc = sum(
            item["hcbdg"] + item["hccrb"] + item["hctsk"]
            for item in rjbb
        )
        totaltur = sum(
            item["turbdg"] + item["turcrb"] + item["turtsk"]
            for item in rjbb
        )
        formatted_totaltur = "{:02.0f}".format(totaltur)

        print(formatted_totaltur)

        cursor.execute("SELECT bulan_hcvstarget, akumulasihc, target, pencapaian FROM tb_retail")
        homepass = cursor.fetchall()

        cursor.close()
        mydb.close()

        return render_template(
            "dashboard/user/dashboard_assets.html", olt=olt,totalhp=totalhp,totalhc=totalhc,formatted_totaltur=formatted_totaltur, totalSum=totalSum, homepass=homepass, totaloltbdg_sum=totaloltbdg_sum,totaloltcrb_sum=totaloltcrb_sum,totalolttsk_sum=totalolttsk_sum,totalportbdg_sum=totalportbdg_sum,totalportcrb_sum=totalportcrb_sum,totalporttsk_sum=totalporttsk_sum,idleportbdg_sum=idleportbdg_sum,idleportcrb_sum=idleportcrb_sum,idleporttsk_sum=idleporttsk_sum,
        )
    except Exception as e:
        return str(e), 500
    
    


@app.route('/process_option', methods=['POST'])
def process_option():
    selected_option = request.form.get('option')

    if selected_option == 'tb_tk':
        return redirect(url_for('tb_tk'))
    
    elif selected_option == 'tb_tur':
        return redirect(url_for('tb_tur'))
    
    
    elif selected_option == 'tb_acum':
        return redirect(url_for('tb_acum'))
    
    elif selected_option == 'tb_acum':
        return redirect(url_for('tb_acum'))

    elif selected_option == 'tb_olt':
        return redirect(url_for('tb_olt'))
    
    elif selected_option == 'tb_weeklyhome':
        return redirect(url_for('tb_weeklyhome'))

    elif selected_option == 'tb_dashboardutama':
        return redirect(url_for('tb_dashboardutama'))
    
    elif selected_option == 'tb_dailyhcspa':
        return redirect(url_for('tb_dailyhcspa'))
    
    elif selected_option == 'tb_jbbmilestone':
        return redirect(url_for('tb_jbbmilestone'))
    
    elif selected_option == 'tb_checkretailassets':
        return redirect(url_for('tb_checkretailassets'))
    
    elif selected_option == 'tb_retailtgl':
        return redirect(url_for('tb_retailtgl'))
    
    elif selected_option == 'tb_aging':
        return redirect(url_for('tb_aging'))
    
    elif selected_option == 'tb_retail':
        return redirect(url_for('tb_retail'))
    
    elif selected_option == 'tb_retail':
        return redirect(url_for('tb_retail'))
    
    elif selected_option == 'ikr':
        return redirect(url_for('ikr'))
    
    elif selected_option == 'pdtfat':
        return redirect(url_for('pdtfat'))
    elif selected_option == 'datamap':
        return redirect(url_for('datamap'))
    else:
        return "Invalid option"
    
#  --------------------tbtur----------------------------- #  
@app.route("/tb_dashboardutama")
def tb_dashboardutama():
    # Get the page number from the request, default to 1 if not provided
    page = request.args.get('page', 1, type=int)
    per_page = 100  # Adjust this number as needed

    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    offset = (page - 1) * per_page
    cursor.execute(f"SELECT * FROM tb_akv LIMIT {per_page} OFFSET {offset}")
    u = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_retail/tb_dashboardutama.html", u=u)


@app.route("/delete_dashboardutama/<int:id>", methods=['POST'])
def delete_dashboardutama(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM tb_akv WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_dashboardutama'))

# New route to delete all data
@app.route("/delete_all_utama", methods=['POST'])
def delete_all_utama():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM tb_akv")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_dashboardutama'))



@app.route("/export_tb_dashboardutama", methods=['GET'])
def export_tb_dashboardutama():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT monthly_bulan, monthly_spa, mothlyhc_bulan, monthly_hc, akumulasihcvsspa_bulan, akumulasihcvsspa_hc, akumulasihcvsspa_target, akumulasihcvsspa_capaian, monthlyhc_bulan, monthlyhc_bdg, monthlyhc_crb, monthlyhc_tsk, akumulasihckp_bulan, akumulasihckp_bdg, akumulasihckp_crb, akumulasihckp_tsk, agingkp_bulan, agingkp_bdg, agingkp_crb, agingkp_tsk, agingkpbai_bulan, agingkpbai_averageaging FROM tb_akv")
    dashboardutama = cursor.fetchall()
    cursor.close()

    # Create a Pandas DataFrame from the database query
    df = pd.DataFrame(dashboardutama)

    # Create an in-memory Excel file
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)

    # Set up response headers for Excel file download
    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=tb_retail_dashboard_utama_export.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

@app.route("/import_tb_dashboardutama", methods=['POST'])
def import_tb_dashboardutama():
    file = request.files['file']
    if file:
        # Read the Excel file into a Pandas DataFrame
        df = pd.read_excel(file)

        # Ensure that the DataFrame columns match the expected columns
        expected_columns = ["monthly_bulan","monthly_spa","mothlyhc_bulan","monthly_hc","akumulasihcvsspa_bulan","akumulasihcvsspa_hc","akumulasihcvsspa_target","akumulasihcvsspa_capaian","monthlyhc_bulan","monthlyhc_bdg","monthlyhc_crb","monthlyhc_tsk","akumulasihckp_bulan","akumulasihckp_bdg","akumulasihckp_crb","akumulasihckp_tsk","agingkp_bulan","agingkp_bdg","agingkp_crb","agingkp_tsk","agingkpbai_bulan","agingkpbai_averageaging"]

        if not all(column in df.columns for column in expected_columns):
            return "Error: Columns in the uploaded file do not match the expected columns."

        # Handle NaN values in the DataFrame by replacing them with None
        df = df.where(pd.notna(df), None)

        # Convert int32 and int64 values to Python Decimal for better MySQL compatibility
        for col in expected_columns:
            if np.issubdtype(df[col].dtype, np.integer):
                df[col] = df[col].astype(int).apply(lambda x: Decimal(x) if x is not None else None)

        # Convert float64 values to Python Decimal for better MySQL compatibility
        for col in expected_columns:
            if df[col].dtype == np.float64:
                df[col] = df[col].apply(lambda x: Decimal(x) if not pd.isna(x) else None)

        # You can now process the DataFrame as needed, for example, insert into the database
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Assuming 'tb_dailyhcspa' is the table name
        # Using executemany to insert multiple rows at once
        columns_str = ', '.join(expected_columns)
        placeholders = ', '.join(['%s' for _ in expected_columns])
        query = f"INSERT INTO tb_akv ({columns_str}) VALUES ({placeholders})"
        
        try:
            # Convert DataFrame values to a list of tuples
            data_to_insert = [tuple(row) for row in df.values]

            cursor.executemany(query, data_to_insert)
            mydb.commit()
            return "File uploaded and data imported successfully."
        except Exception as e:
            return f"Error: {e}"
        finally:
            cursor.close()

    return "No file provided."






@app.route("/tb_dailyhcspa")
def tb_dailyhcspa():
    # Get the page number from the request, default to 1 if not provided
    page = request.args.get('page', 1, type=int)
    per_page = 100  # Adjust this number as needed

    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    offset = (page - 1) * per_page
    cursor.execute(f"SELECT * FROM tb_dailyhcspa LIMIT {per_page} OFFSET {offset}")
    daily = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_retail/tb_dailyhc_spa.html", daily=daily)

@app.route("/delete_dailyhcspa/<int:id>", methods=['POST'])
def delete_dailyhcspa(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM tb_dailyhcspa WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_dailyhcspa'))

# New route to delete all data
@app.route("/delete_all_dailyhcspa", methods=['POST'])
def delete_all_dailyhcspa():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM tb_dailyhcspa")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_dailyhcspa'))

@app.route("/export_tb_dailyhcspa", methods=['GET'])
def export_tb_dailyhcspa():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT tgl, qty, tglhc, hc, tglspabdg, qtyspabdg, tglhcbdg, hc_hcbdg, tglspacrb, qtyspacrb, tglhccrb, hc_hccrb, tglspatsk, qtyspatsk, tglhctsk, hc_hctsk FROM tb_dailyhcspa")
    dailyhcspa = cursor.fetchall()
    cursor.close()

    # Create a Pandas DataFrame from the database query
    df = pd.DataFrame(dailyhcspa)

    # Create an in-memory Excel file
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)

    # Set up response headers for Excel file download
    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=tb_dailyhcspa_export.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route("/import_tb_dailyhcspa", methods=['POST'])
def import_tb_dailyhcspa():
    file = request.files['file']
    if file:
        # Read the Excel file into a Pandas DataFrame
        df = pd.read_excel(file)

        # Ensure that the DataFrame columns match the expected columns
        expected_columns = ["tgl", "qty", "tglhc", "hc", "tglspabdg", "qtyspabdg", "tglhcbdg", "hc_hcbdg",
                            "tglspacrb", "qtyspacrb", "tglhccrb", "hc_hccrb", "tglspatsk", "qtyspatsk", "tglhctsk", "hc_hctsk"]

        if not all(column in df.columns for column in expected_columns):
            return "Error: Columns in the uploaded file do not match the expected columns."

        # Handle NaN values in the DataFrame by replacing them with None
        df = df.where(pd.notna(df), None)

        # Convert int32 and int64 values to Python Decimal for better MySQL compatibility
        for col in expected_columns:
            if np.issubdtype(df[col].dtype, np.integer):
                df[col] = df[col].astype(int).apply(lambda x: Decimal(x) if x is not None else None)

        # Convert float64 values to Python Decimal for better MySQL compatibility
        for col in expected_columns:
            if df[col].dtype == np.float64:
                df[col] = df[col].apply(lambda x: Decimal(x) if not pd.isna(x) else None)

        # You can now process the DataFrame as needed, for example, insert into the database
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Assuming 'tb_dailyhcspa' is the table name
        # Using executemany to insert multiple rows at once
        columns_str = ', '.join(expected_columns)
        placeholders = ', '.join(['%s' for _ in expected_columns])
        query = f"INSERT INTO tb_dailyhcspa ({columns_str}) VALUES ({placeholders})"
        
        try:
            # Convert DataFrame values to a list of tuples
            data_to_insert = [tuple(row) for row in df.values]

            cursor.executemany(query, data_to_insert)
            mydb.commit()
            return "File uploaded and data imported successfully."
        except Exception as e:
            return f"Error: {e}"
        finally:
            cursor.close()

    return "No file provided."


@app.route("/tb_jbbmilestone")
def tb_jbbmilestone():
    # Get the page number from the request, default to 1 if not provided
    page = request.args.get('page', 1, type=int)
    per_page = 100  # Adjust this number as needed

    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    offset = (page - 1) * per_page
    cursor.execute(f"SELECT * FROM tb_jbbmilestone LIMIT {per_page} OFFSET {offset}")
    stone = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_retail/tb_jbbmilestone.html", stone=stone)

@app.route("/delete_jbbmilestone/<int:id>", methods=['POST'])
def delete_jbbmilestone(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM tb_jbbmilestone WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_jbbmilestone'))

# New route to delete all data
@app.route("/delete_all_jbbmilestone", methods=['POST'])
def delete_all_jbbmilestone():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM tb_jbbmilestone")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_jbbmilestone'))


@app.route("/export_tb_jbbmilestone", methods=['GET'])
def export_tb_jbbmilestone():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tb_jbbmilestone")
    jbbmilestone = cursor.fetchall()
    cursor.close()

    # Create a Pandas DataFrame from the database query
    df = pd.DataFrame(jbbmilestone)

    # Create an in-memory Excel file
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)

    # Set up response headers for Excel file download
    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=tb_jbbmilestone_export.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

@app.route("/import_tb_jbbmilestone", methods=['POST'])
def import_tb_jbbmilestone():
    file = request.files['file']
    if file:
        # Read the Excel file into a Pandas DataFrame
        df = pd.read_excel(file)

        # Ensure that the DataFrame columns match the expected columns
        expected_columns = ["tahun", "bandung", "cirebon", "tasik"]

        if not all(column in df.columns for column in expected_columns):
            return "Error: Columns in the uploaded file do not match the expected columns."

        # Handle NaN values in the DataFrame by replacing them with None
        df = df.applymap(lambda x: None if pd.isna(x) else x)

        # You can now process the DataFrame as needed, for example, insert into the database
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Assuming 'tb_jbbmilestone' is the table name
        # Using executemany to insert multiple rows at once
        columns_str = ', '.join(expected_columns)
        placeholders = ', '.join(['%s' for _ in expected_columns])
        query = f"INSERT INTO tb_jbbmilestone ({columns_str}) VALUES ({placeholders})"
        
        try:
            cursor.executemany(query, df.values.tolist())
            mydb.commit()
            return "File uploaded and data imported successfully."
        except Exception as e:
            return f"Error: {e}"

        finally:
            cursor.close()

    return "No file provided."




@app.route("/tb_checkretailassets")
def tb_checkretailassets():
    # Get the page number from the request, default to 1 if not provided
    page = request.args.get('page', 1, type=int)
    per_page = 100  # Adjust this number as needed

    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    offset = (page - 1) * per_page
    cursor.execute(f"SELECT * FROM tb_checkretail LIMIT {per_page} OFFSET {offset}")
    retail = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_retail/tb_checkretail.html", retail=retail)



@app.route("/delete_checkretailassets/<int:id>", methods=['POST'])
def delete_checkretailassets(id):
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Your logic to delete the specified row
    delete_query = "DELETE FROM tb_checkretail WHERE id = %s"
    cursor.execute(delete_query, (id,))
    
    mydb.commit()
    cursor.close()
    mydb.close()

    # Redirect back to the checkretail assets page
    return redirect(url_for('tb_checkretailassets'))


# New route to delete all data
@app.route("/delete_all_checkretailassets", methods=['POST'])
def delete_all_checkretailassets():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM tb_checkretail")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_checkretailassets'))

@app.route("/export_tb_checkretail", methods=['GET'])
def export_tb_checkretail():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT tgl, hc, rev, tglbdg, hcbdg, revbdg, tglcrb, hccrb, revcrb, tgltsk, hctsk, revtsk FROM tb_checkretail")
    checkretail = cursor.fetchall()
    cursor.close()

    # Create a Pandas DataFrame from the database query
    df = pd.DataFrame(checkretail)

    # Create an in-memory Excel file
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)

    # Set up response headers for Excel file download
    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=tb_checkretail_export.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

@app.route("/import_tb_checkretail", methods=['POST'])
def import_tb_checkretail():
    file = request.files['file']
    if file:
        # Read the Excel file into a Pandas DataFrame
        df = pd.read_excel(file)

        # Ensure that the DataFrame columns match the expected columns
        expected_columns = ["tgl", "hc", "rev", "tglbdg", "hcbdg", "revbdg", "tglcrb", "hccrb", "revcrb", "tgltsk", "hctsk", "revtsk"]

        if not all(column in df.columns for column in expected_columns):
            return "Error: Columns in the uploaded file do not match the expected columns."

        # Handle NaN values in the DataFrame by replacing them with None
        df = df.applymap(lambda x: None if pd.isna(x) else x)

        # You can now process the DataFrame as needed, for example, insert into the database
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Assuming 'tb_checkretail' is the table name
        # Using executemany to insert multiple rows at once
        columns_str = ', '.join(expected_columns)
        placeholders = ', '.join(['%s' for _ in expected_columns])
        query = f"INSERT INTO tb_checkretail ({columns_str}) VALUES ({placeholders})"
        
        try:
            cursor.executemany(query, df.values.tolist())
            mydb.commit()
            return "File uploaded and data imported successfully."
        except Exception as e:
            return f"Error: {e}"

        finally:
            cursor.close()

    return "No file provided."




@app.route("/tb_retailtgl")
def tb_retailtgl():
    # Get the page number from the request, default to 1 if not provided
    page = request.args.get('page', 1, type=int)
    per_page = 100  # Adjust this number as needed

    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    offset = (page - 1) * per_page
    cursor.execute(f"SELECT * FROM tb_retail LIMIT {per_page} OFFSET {offset}")
    retaily = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_retail/tb_retail.html", retaily=retaily)

@app.route("/delete_retailtgl/<int:id>", methods=['POST'])
def delete_retailtgl(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM tb_retail WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_retailtgl'))

# New route to delete all data
@app.route("/delete_all_retailtgl", methods=['POST'])
def delete_all_retailtgl():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM tb_retail")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_retailtgl'))

@app.route("/export_tb_retail", methods=['GET'])
def export_tb_retail():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tb_retail")
    retaily = cursor.fetchall()
    cursor.close()

    # Create a Pandas DataFrame from the database query
    df = pd.DataFrame(retaily)

    # Exclude the 'id' column from the DataFrame
    df = df[['bulan', 'rev', 'accumrev', 'bulanbdg', 'revbdg', 'accum_revbdg', 'bulancrb', 'revcrb', 'accum_revcrb',
             'bulantsk', 'revtsk', 'accum_revtsk', 'montly_rev_bulan', 'montly_rev_hc', 'montly_rev', 'accumm_rev',
             'bulan_badwhith', 'hc_badwhith', 'bulan_hcvstarget', 'akumulasihc', 'target', 'pencapaian']]

    # Create an in-memory Excel file
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)

    # Set up response headers for Excel file download
    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=tb_retailtgl_export.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response



@app.route("/import_tb_retail", methods=['POST'])
def import_tb_retail():
    file = request.files['file']
    if file:
        # Read the Excel file into a Pandas DataFrame
        df = pd.read_excel(file)

        # Ensure that the DataFrame columns match the expected columns
        expected_columns = ['bulan', 'rev', 'accumrev', 'bulanbdg', 'revbdg', 'accum_revbdg', 'bulancrb', 'revcrb',
                            'accum_revcrb', 'bulantsk', 'revtsk', 'accum_revtsk', 'montly_rev_bulan', 'montly_rev_hc',
                            'montly_rev', 'accumm_rev', 'bulan_badwhith', 'hc_badwhith', 'bulan_hcvstarget', 'akumulasihc',
                            'target', 'pencapaian']

        if not all(column in df.columns for column in expected_columns):
            return "Error: Columns in the uploaded file do not match the expected columns."

        # Handle NaN values in the DataFrame by replacing them with None
        df = df.applymap(lambda x: None if pd.isna(x) else x)

        # You can now process the DataFrame as needed, for example, insert into the database
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Assuming 'tb_retail' is the table name
        # Using executemany to insert multiple rows at once
        columns_str = ', '.join(expected_columns)
        placeholders = ', '.join(['%s' for _ in expected_columns])
        query = f"INSERT INTO tb_retail ({columns_str}) VALUES ({placeholders})"
        
        try:
            cursor.executemany(query, df.values.tolist())
            mydb.commit()
            return "File uploaded and data imported successfully."
        except Exception as e:
            return f"Error: {e}"

        finally:
            cursor.close()

    return "No file provided."

# ... (remaining code)

# ... (remaining code)



@app.route("/tb_aging")
def tb_aging():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT * FROM jbbaging")
    aging = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_retail/tb_aging.html", aging=aging)

@app.route("/delete_aging/<int:id>", methods=['POST'])
def delete_aging(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM jbbaging WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_aging'))

# New route to delete all data
@app.route("/delete_all_aging", methods=['POST'])
def delete_all_aging():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM jbbaging ")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_aging'))


@app.route("/export_tb_jbbaging")
def export_tb_jbbaging():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT kabagingbdg, kpibdg, agingbdg, kabagingcrb, kpicrb, agingcrb, kabagingtsk, kpitsk, agingtsk, agingikpbaseonbai FROM jbbaging")
    aging_data = cursor.fetchall()
    cursor.close()

    df = pd.DataFrame(aging_data)

    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)
    excel_data.seek(0)

    return Response(excel_data.getvalue(),
                    headers={'Content-Disposition': 'attachment; filename=tb_jbbaging_export.xlsx'},
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route("/import_tb_jbbaging", methods=['POST'])
def import_tb_jbbaging():
    file = request.files['file']
    if file:
        # Read the Excel file into a Pandas DataFrame
        df = pd.read_excel(file)

        # Ensure that the DataFrame columns match the expected columns
        expected_columns = ["kabagingbdg", "kpibdg", "agingbdg", "kabagingcrb", "kpicrb", "agingcrb", "kabagingtsk", "kpitsk", "agingtsk", "agingikpbaseonbai"]

        if not all(column in df.columns for column in expected_columns):
            return "Error: Columns in the uploaded file do not match the expected columns."

        # Handle NaN values in the DataFrame by replacing them with None
        df = df.where(pd.notna(df), None)

        # You can now process the DataFrame as needed, for example, insert into the database
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Assuming 'tb_jbbaging' is the table name
        # Using executemany to insert multiple rows at once
        columns_str = ', '.join(expected_columns)
        placeholders = ', '.join(['%s' for _ in expected_columns])
        query = f"INSERT INTO jbbaging ({columns_str}) VALUES ({placeholders})"
        
        try:
            cursor.executemany(query, df.values.tolist())
            mydb.commit()
            return "File uploaded and data imported successfully."
        except Exception as e:
            return f"Error: {e}"
        finally:
            cursor.close()

    return "No file provided."

@app.route("/tb_tur")
def tb_tur():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT * FROM tb_tur")
    tur = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_assets/tb_tur.html", tur=tur)

@app.route("/delete_tur/<int:id>", methods=['POST'])
def delete_tur(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM tb_tur WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_tur'))

# New route to delete all data
@app.route("/delete_all_tb_tur", methods=['POST'])
def delete_all_tb_tur():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM tb_tur ")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_tur'))

@app.route('/export_tb_tur_excel')
def export_tb_tur_excel():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT * FROM tb_tur")
    tur = cursor.fetchall()

    cursor.close()

    excel_directory = "uploads/tb_tur/"
    os.makedirs(excel_directory, exist_ok=True)

    # Create an Excel file
    excel_filename = os.path.join(excel_directory, "tb_tur_export.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write header row
    header = ["kabbdg", "hpbdg", "hcbdg", "turbdg", "kabcrb", "hpcrb", "hccrb", "turcrb", "kabtsk", "hptsk", "hctsk", "turtsk"]
    ws.append(header)

    # Write data rows
    for row in tur:
        ws.append([row[col] for col in header])

    # Save the Excel file
    wb.save(excel_filename)

    # Move the Excel file to the desired directory
    shutil.move(excel_filename, excel_directory)

    # Provide the Excel file as a response
    return Response(
        open(os.path.join(excel_directory, "tb_tur_export.xlsx"), 'rb'),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename=tb_tur_export.xlsx"}
    )


@app.route('/export_tb_tur_csv', methods=['GET'])
def export_tb_tur_csv():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    try:
        # Fetch data from the database
        cursor.execute("SELECT * FROM tb_tur")
        rows = cursor.fetchall()

        # Specify the file name for the CSV export
        csv_directory = "uploads/tb_tur_csv"
        os.makedirs(csv_directory, exist_ok=True)
        csv_filename = os.path.join(csv_directory, "tb_tur_export.csv")

        # Write data to the CSV file
        with open(csv_filename, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)
            
            # Write header row
            header_row = ["kabbdg", "hpbdg", "hcbdg", "turbdg", "kabcrb", "hpcrb", "hccrb", "turcrb", "kabtsk", "hptsk", "hctsk", "turtsk"]
            csv_writer.writerow(header_row)
            
            # Write data rows
            for row in rows:
                csv_writer.writerow([row[key] for key in header_row])

        # Provide the CSV file for download
        return send_file(csv_filename, as_attachment=True)
    except Exception as e:
        print(f"Error exporting CSV: {e}")
    finally:
        cursor.close()
        mydb.close()

    # Handle other cases, e.g., return to the original page
    return redirect(url_for('tb_tur'))

# Your existing route
@app.route('/import_tb_tur_excel', methods=['POST'])
def import_tb_tur_excel():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    if request.method == 'POST':
        file = request.files['file']

        if file:
            # Save the uploaded Excel file
            file.save(file.filename)

            try:
                # Read the Excel file and insert data into the database
                wb = openpyxl.load_workbook(file.filename)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Ensure the number of columns in the row matches the number of placeholders in the query
                    if len(row) == 12:  # Adjust the number based on your actual column count
                        query = """
                            INSERT INTO tb_tur (
                                kabbdg, hpbdg, hcbdg, turbdg, kabcrb, hpcrb, hccrb, turcrb, kabtsk, hptsk, hctsk, turtsk
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """
                        cursor.execute(query, row)
                        mydb.commit()
                    else:
                        print(f"Invalid row data: {row}")

                # Redirect to the original page
                return redirect(url_for('tb_tur'))
            except Exception as e:
                print(f"Error reading Excel file: {e}")

    # Handle other cases, e.g., GET request
    return redirect(url_for('tb_tur'))
    

@app.route('/import_tb_tur_csv', methods=['POST'])
def import_tb_tur_csv():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    if request.method == 'POST':
        file = request.files['file']

        if file and file.filename.endswith('.csv'):
            # Save the uploaded CSV file
            file.save(file.filename)

            try:
                # Read the CSV file and insert data into the database
                with open(file.filename, 'r', newline='') as csv_file:
                    csv_reader = csv.reader(csv_file, delimiter=';')
                    next(csv_reader)  # Skip the header row if it exists

                    for row in csv_reader:
                        # Split each row into individual values
                        values = []
                        for item in row:
                            values.extend(item.split(';'))

                        # Ensure the number of columns in the values list matches the number of placeholders in the query
                        if len(values) == 12:  # Adjust the number based on your actual column count
                            query = """
                                INSERT INTO tb_tur (
                                    kabbdg, hpbdg, hcbdg, turbdg, kabcrb, hpcrb, hccrb, turcrb, kabtsk, hptsk, hctsk, turtsk
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """
                            cursor.execute(query, values)
                            mydb.commit()
                        else:
                            print(f"Invalid row data: {row}")

                # Redirect to the original page
                return redirect(url_for('tb_tur'))
            except Exception as e:
                print(f"Error reading CSV file: {e}")

    # Handle other cases, e.g., GET request
    return redirect(url_for('tb_tur'))


# -----------------------endtb_tur-----------------------#

# -------------------------tbtk-------------------------#
@app.route("/tb_tk")
def tb_tk():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT * FROM  tb_akv_assets")
    tk = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_assets/tb_tk.html", tk=tk)

@app.route("/delete_tk/<int:id>", methods=['POST'])
def delete_tk(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM tb_akv_assets WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_tk'))

# New route to delete all data
@app.route("/delete_all_tb_tk", methods=['POST'])
def delete_all_tb_tk():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM tb_akv_assets ")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_tk'))


@app.route('/export_tb_tk_excel')
def export_tb_tk_excel():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT * FROM tb_akv_assets")
    tk = cursor.fetchall()

    cursor.close()

    excel_directory = "uploads/tb_tk/"
    os.makedirs(excel_directory, exist_ok=True)

    # Create an Excel file
    excel_filename = os.path.join(excel_directory, "tb_tk_export.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write header row
    header = ["kabtk1", "captk1", "hptk1", "hctk1", "sumofturtk1", "kabtk2", "captk2", "hptk2", "hctk2", "sumofturtk2", "kabtk4", "captk4", "hptk4", "hctk4", "sumofturtk4", "kab_invest", "captk_invest", "hptk_invest", "hctk_invest", "sumofturtk_invest", "kabcb1", "capcb1", "hpcb1", "hccb1", "sumofturcb1", "kabcb2", "capcb2", "hpcb2", "hccb2", "sumofturcb2", "kabcb3", "capcb3", "hpcb3", "hccb3", "sumofturcb3"]
    ws.append(header)

    # Write data rows
    for row in tk:
        ws.append([row[col] for col in header])

    # Save the Excel file
    wb.save(excel_filename)

    # Move the Excel file to the desired directory
    shutil.move(excel_filename, excel_directory)

    # Provide the Excel file as a response
    return Response(
        open(os.path.join(excel_directory, "tb_tk_export.xlsx"), 'rb'),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename=tb_tk_export.xlsx"}
    )

# Your existing route
@app.route('/import_tb_tk_excel', methods=['POST'])
def import_tb_tk_excel():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    if request.method == 'POST':
        file = request.files['file']

        if file:
            # Save the uploaded Excel file
            file.save(file.filename)

            try:
                # Read the Excel file and insert data into the database
                wb = openpyxl.load_workbook(file.filename)
                ws = wb.active
                column_count = 35  # Adjust the number based on your actual column count

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # Ensure the number of columns in the row matches the number of placeholders in the query
                    if len(row) != column_count:
                        print(f"Invalid row data at row {row_num}: {row}")
                        continue

                    # Process values and replace commas with periods for decimal numbers
                    processed_row = []
                    for cell in row:
                        if isinstance(cell, (int, float)):
                            # Preserve the original format of numeric values
                            processed_row.append(str(cell).replace(',', '.'))
                        else:
                            # Preserve the original format of other values
                            processed_row.append(cell)

                    # Fill in default values (0) for empty cells
                    processed_row = [0 if cell is None else cell for cell in processed_row]

                    query = f"""
                            INSERT INTO tb_akv_assets (kabtk1, captk1, hptk1, hctk1, sumofturtk1, kabtk2, captk2, hptk2, hctk2, sumofturtk2,
                            kabtk4, captk4, hptk4, hctk4, sumofturtk4, kab_invest, captk_invest, hptk_invest, hctk_invest, sumofturtk_invest,
                            kabcb1, capcb1, hpcb1, hccb1, sumofturcb1, kabcb2, capcb2, hpcb2, hccb2, sumofturcb2,
                            kabcb3, capcb3, hpcb3, hccb3, sumofturcb3)
                            VALUES ({', '.join(['%s'] * column_count)});
                            """
                    cursor.execute(query, processed_row)
                    mydb.commit()

                # Redirect to the original page
                return redirect(url_for('tb_tk'))
            except Exception as e:
                print(f"Error reading or inserting Excel file data: {e}")

    # Handle other cases, e.g., GET request
    return redirect(url_for('tb_tk'))










@app.route("/tb_olt")
def tb_olt():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT * FROM tb_olt")
    olt = cursor.fetchall()

    cursor.close()
    return render_template("tb_admin/tb_assets/tb_olt.html", olt=olt)
# New route to delete a single row

@app.route("/delete_olt/<int:id>", methods=['POST'])
def delete_olt(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM tb_olt WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_olt'))

# New route to delete all data
@app.route("/delete_all_olt", methods=['POST'])
def delete_all_olt():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM tb_olt")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_olt'))


@app.route('/export_tb_olt_excel')
def export_tb_olt_excel():
    try:
        mydb = get_db_connection()
        cursor = mydb.cursor(dictionary=True)

        cursor.execute("SELECT * FROM tb_olt")
        olt_data = cursor.fetchall()

        cursor.close()

        # Directory for saving Excel file
        excel_directory = "uploads/tb_olt/"
        os.makedirs(excel_directory, exist_ok=True)

        # Create an Excel file
        excel_filename = os.path.join(excel_directory, "tb_olt_export.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write header row
        header = ["totaloltbdg", "totalportbdg", "idleportbdg", "totaloltcrb", "totalportcrb", "idleportcrb", "totalolttsk", "totalporttsk", "idleporttsk"]
        ws.append(header)

        # Write data rows
        for row in olt_data:
            ws.append([row[col] for col in header])

        # Save the Excel file
        wb.save(excel_filename)

        # Provide the Excel file as a response
        response = Response(
            open(excel_filename, 'rb'),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": f"attachment; filename=tb_olt_export.xlsx"}
        )

        return response

    except Exception as e:
        print(f"Error exporting to Excel file: {e}")
        # Handle the error, e.g., return an error response

    # In case of an error, you might want to return an appropriate error response
    return "Error exporting to Excel file"


@app.route('/import_tb_olt_excel', methods=['POST'])
def import_tb_olt_excel():
    try:
        mydb = get_db_connection()
        cursor = mydb.cursor(dictionary=True)

        if request.method == 'POST':
            file = request.files['file']

            if file:
                # Save the uploaded Excel file
                file.save(file.filename)

                # Read the Excel file and insert/update data into the database
                wb = openpyxl.load_workbook(file.filename)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) == 9:  # Adjust the number based on your actual column count
                        query = """
                            INSERT INTO tb_olt (totaloltbdg, totalportbdg, idleportbdg, totaloltcrb, totalportcrb, idleportcrb, totalolttsk, totalporttsk, idleporttsk)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                            totaloltbdg = COALESCE(%s, totaloltbdg),
                            totalportbdg = COALESCE(%s, totalportbdg),
                            idleportbdg = COALESCE(%s, idleportbdg),
                            totaloltcrb = COALESCE(%s, totaloltcrb),
                            totalportcrb = COALESCE(%s, totalportcrb),
                            idleportcrb = COALESCE(%s, idleportcrb),
                            totalolttsk = COALESCE(%s, totalolttsk),
                            totalporttsk = COALESCE(%s, totalporttsk),
                            idleporttsk = COALESCE(%s, idleporttsk);
                        """
                        # Duplicate each value to match the ON DUPLICATE KEY UPDATE clause
                        params = row + row

                        cursor.execute(query, params)
                        mydb.commit()
                    else:
                        print(f"Invalid row data: {row}")

                # Redirect to the original page
                return redirect(url_for('tb_olt'))

    except Exception as e:
        print(f"Error reading Excel file: {e}")

    # Handle other cases, e.g., GET request or error
    return redirect(url_for('tb_olt'))



@app.route("/tb_acum")
def tb_acum():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT * FROM acumhompass")
    acum = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_assets/tb_acum.html", acum=acum)

@app.route("/delete_acum/<int:id>", methods=['POST'])
def delete_acum(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM acumhompass WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_acum'))

# New route to delete all data
@app.route("/delete_all_acum", methods=['POST'])
def delete_all_acum():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM acumhompass")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_acum has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_acum'))


@app.route('/export_tb_akumulasi_excel')
def export_tb_akumulasi_excel():
    try:
        mydb = get_db_connection()
        cursor = mydb.cursor(dictionary=True)

        cursor.execute("SELECT * FROM acumhompass")
        acum = cursor.fetchall()

        cursor.close()

        # Directory for saving Excel file
        excel_directory = "uploads/tb_akumulasi/"
        os.makedirs(excel_directory, exist_ok=True)

        # Create an Excel file
        excel_filename = os.path.join(excel_directory, "tb_akumulasi_export.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write header row
        header = ["bulan", "akumulasi", "target", "pencapaian"]
        ws.append(header)

        # Write data rows
        for row in acum:
            ws.append([row[col] for col in header])

        # Save the Excel file
        wb.save(excel_filename)

        # Provide the Excel file as a response
        response = Response(
            open(excel_filename, 'rb'),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": f"attachment; filename=tb_akumulasi_export.xlsx"}
        )

        return response

    except Exception as e:
        print(f"Error exporting to Excel file: {e}")
        # Handle the error, e.g., return an error response

    # In case of an error, you might want to return an appropriate error response
    return "Error exporting to Excel file"



# Import data dari file Excel (xlsx) ke tabel acumhompass
@app.route('/import_tb_acum', methods=['POST'])
def import_tb_acum():
    if 'file' not in request.files:
        flash('File tidak ditemukan', 'danger')
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        flash('Tidak ada file yang dipilih', 'danger')
        return redirect(request.url)

    if file:
        try:
            df = pd.read_excel(file)
            acum_data = df.to_dict('records')

            mydb = get_db_connection()
            cursor = mydb.cursor(dictionary=True)

            for data in acum_data:
                columns = ', '.join(data.keys())
                values = ', '.join(['%s'] * len(data))
                query = f"INSERT INTO acumhompass ({columns}) VALUES ({values})"
                cursor.execute(query, tuple(data.values()))

            mydb.commit()
            flash('Data berhasil diimpor ke acumhompass', 'success')
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
        finally:
            cursor.close()
            mydb.close()

    return redirect(url_for('tb_acum'))


@app.route("/tb_weeklyhome")
def tb_weeklyhome():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT * FROM weeklyhompass ")
    home = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_assets/tb_weeklyhome.html", home=home)

@app.route("/delete_weeklyhome/<int:id>", methods=['POST'])
def delete_weeklyhome(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM weeklyhompass WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_weeklyhome'))

# New route to delete all data
@app.route("/delete_all_tb_weeklyhompass", methods=['POST'])
def delete_all_tb_weeklyhompass():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM weeklyhompass ")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('tb_weeklyhome'))


@app.route('/export_tb_weeklyhome')
def export_tb_weeklyhome():
    try:
        mydb = get_db_connection()
        cursor = mydb.cursor(dictionary=True)

        cursor.execute("SELECT * FROM weeklyhompass")
        home = cursor.fetchall()

        cursor.close()

        # Directory for saving Excel file
        excel_directory = "uploads/tb_weeklyhome/"
        os.makedirs(excel_directory, exist_ok=True)

        # Create an Excel file
        excel_filename = os.path.join(excel_directory, "tb_weeklyhome_export.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write header row
        header = ["bulan_tgl", "sumofkafasitas"]
        ws.append(header)

        # Write data rows
        for row in home:
            ws.append([row[col] for col in header])

        # Save the Excel file
        wb.save(excel_filename)

        # Provide the Excel file as a response
        response = Response(
            open(excel_filename, 'rb'),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-disposition": f"attachment; filename=tb_weeklyhome_export.xlsx"}
        )
        return response

    except Exception as e:
        print(f"Error exporting to Excel file: {e}")
        # Handle the error, e.g., return an error response

    # In case of an error, you might want to return an appropriate error response
    return "Error exporting to Excel file"


# Import data dari file Excel (xlsx) ke tabel acumhompass
@app.route('/import_tb_weeklyhome', methods=['POST'])
def import_tb_weeklyhome():
    if 'file' not in request.files:
        flash('File tidak ditemukan', 'danger')
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        flash('Tidak ada file yang dipilih', 'danger')
        return redirect(request.url)

    if file:
        try:
            df = pd.read_excel(file)
            home= df.to_dict('records')

            mydb = get_db_connection()
            cursor = mydb.cursor(dictionary=True)

            for data in home:
                columns = ', '.join(data.keys())
                values = ', '.join(['%s'] * len(data))
                query = f"INSERT INTO weeklyhompass ({columns}) VALUES ({values})"
                cursor.execute(query, tuple(data.values()))

            mydb.commit()
            flash('Data berhasil diimpor ke weeklyhompass', 'success')
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
        finally:
            cursor.close()
            mydb.close()

    return redirect(url_for('tb_weeklyhome'))



@app.route("/preesale")
def preesale():
    return render_template("error.html")


@app.route("/gangguan")
def gangguan():
    return render_template("error.html")

@app.route("/ikr", methods=['GET'])
def ikr():
    # Get the page number from the request, default to 1 if not provided
    page = request.args.get('page', 1, type=int)
    per_page = 1000  # Adjust this number as needed

    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    offset = (page - 1) * per_page
    cursor.execute(f"SELECT * FROM  oltdata LIMIT {per_page} OFFSET {offset}")
    ikr = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_assets/dataikr.html",ikr=ikr)

@app.route("/delete_ikr/<int:id>", methods=['POST'])
def delete_ikr(id):
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete a single row from tb_olt
        cursor.execute("DELETE FROM  oltdata WHERE id = %s", (id,))
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("Row deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting the row. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('ikr'))

# New route to delete all data
@app.route("/delete_all_ikr", methods=['POST'])
def delete_all_ikr():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM oltdata ")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('ikr'))


def export_excel(df):
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)
    return excel_data

def import_excel(file):
    df = pd.read_excel(file)
    return df


def save_to_database(df):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        for _, row in df.iterrows():
            # Exclude the first element (primary key) from the data tuple
            data_tuple = tuple(row[1:])
            
            sql_insert = """
                        INSERT INTOoltdata (
                            hostname, pop, area, kota, kecamatan, kelurahan, up3, ulp, asal_olt, jenis_olt, brand_olt, type,
                            kapasitas, port_idle, homepass, home_connected, utilitas, instalation_olt, jumlah_port_olt, nms_olt,
                            l3_switch, status_olt, aging, kategori_cluster, kategori_aging, kategori_tur, idle_hp,
                            jumlah_port_terpakai, koordinat
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s
                        );
                    """

            print(f"Executing SQL: {sql_insert}, Data Tuple: {data_tuple}")

            cursor.execute(sql_insert, data_tuple)

        conn.commit()
    except Exception as e:
        print(f"Error inserting data into the database: {e}")
    finally:
        cursor.close()
        conn.close()


@app.route('/import', methods=['POST'])
def import_data():
    if request.method == 'POST':
        file = request.files['file']
        try:
            df = import_excel(file)
            save_to_database(df)
            return redirect(url_for('ikr'))
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            return redirect(url_for('ikr'))  # Redirect to 'ikr' even if there is an error


@app.route('/export', methods=['GET'])
def export_data():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT hostname,pop,area,kota,kecamatan,kelurahan,up3,ulp,asal_olt,jenis_olt,brand_olt,type,kapasitas,port_idle,homepass,home_connected,utilitas,instalation_olt,jumlah_port_olt,nms_olt,l3_switch,status_olt,aging,kategori_cluster,kategori_aging,kategori_tur,idle_hp,jumlah_port_terpakai,koordinat FROM oltdata")
    data_from_db = cursor.fetchall()
    cursor.close()

    df = pd.DataFrame(data_from_db)
    excel_data = export_excel(df)

    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=tb_oltdata_export.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


# @app.route('/import_tb_ikr_csv', methods=['POST'])
# def import_tb_ikr_csv():
#     mydb = get_db_connection()
#     cursor = mydb.cursor(dictionary=True)

#     if request.method == 'POST':
#         file = request.files['file']

#         if file and file.filename.endswith('.csv'):
#             try:
#                 # Read the CSV file and insert data into the database
#                # Read the CSV file and insert data into the database
#                 with open(file.filename, 'r', newline='') as csv_file:
#                     csv_reader = csv.reader(csv_file, delimiter=';')
#                     header = next(csv_reader, None)  # Get the header row

#                     if header:
#                         # Split the header string into individual column names
#                         header = header[0].split(',')
#                        # Validate header columns against expected columns
#                         expected_columns = [
#                             'hostname', 'pop', 'area', 'kota', 'kecamatan', 'kelurahan', 'up3', 'ulp', 'asal_olt', 'jenis_olt',
#                             'brand_olt', 'type', 'kapasitas', 'port_idle', 'homepass', 'home_connected', 'utilitas', 
#                             'instalation_olt', 'jumlah_port_olt', 'nms_olt', 'l3_switch', 'status_olt', 'aging', 'kategori_cluster',
#                             'kategori_aging', 'kategori_tur', 'idle_hp', 'jumlah_port_terpakai', 'koordinat'
#                         ]

#                         if set(header) != set(expected_columns):
#                             flash('Invalid header columns in the CSV file. Please check the column names and order.', 'error')
                            
#                             # Debug prints for troubleshooting
#                             print("CSV Header:", header)
#                             print("Expected Columns:", expected_columns)
                            
#                             return redirect(url_for('ikr'))

#                         for row_number, row in enumerate(csv_reader, start=2):  # Start counting rows from 2 (header is row 1)
#                             # Split each row into individual values
#                             values = row

#                             # Ensure the number of columns in the values list matches the number of placeholders in the query
#                             if len(values) == len(expected_columns):
#                                 query = """
#                                 INSERT INTO oltdata (
#                                     hostname, pop, area, kota, kecamatan, kelurahan, up3, ulp, asal_olt, jenis_olt, brand_olt, type,
#                                     kapasitas, port_idle, homepass, home_connected, utilitas, instalation_olt, jumlah_port_olt, nms_olt,
#                                     l3_switch, status_olt, aging, kategori_cluster, kategori_aging, kategori_tur, idle_hp,
#                                     jumlah_port_terpakai, koordinat
#                                 ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
#                                 %s, %s, %s, %s, %s, %s)
#                                 """
#                                 try:
#                                     cursor.execute(query, values)
#                                     mydb.commit()
#                                 except Exception as e:
#                                     flash(f"Error inserting data in row {row_number}: {e}", 'error')
#                                     # Debug print for troubleshooting
#                                     print(f"Invalid data in row {row_number}: {values}")
#                             else:
#                                 flash(f'Invalid data in row {row_number}: Incorrect number of columns.', 'error')
#                                 # Debug print for troubleshooting
#                                 print(f"Invalid data in row {row_number}: {values}")


#                 # Redirect to the original page
#                 return redirect(url_for('ikr'))
#             except Exception as e:
#                 flash(f"Error reading CSV file: {e}", 'error')

#     # Handle other cases, e.g., GET request
#     return redirect(url_for('ikr'))




@app.route("/mapikruser")
def mapikruser():
    # Establish a database connection
    mydb = get_db_connection()

    if mydb is None:
        # Handle the case where the connection couldn't be established
        return "Unable to connect to the database", 500

    cursor = mydb.cursor(dictionary=True)

    try:
        cursor.execute("SELECT row_kab_bandung, hp_kab_bandung,hc_kab_bandung, tur_kab_bandung, row_kab_bandungbarat, hp_kab_bandungbarat,hc_kab_bandungbarat,tur_kab_bandungbarat,row_kab_ciamis,hp_kab_ciamis,hc_kab_ciamis,tur_kab_ciamis,row_kab_cianjur,hp_kab_cianjur,hc_kab_cianjur,tur_kab_cianjur,row_kab_cirebon,hp_kab_cirebon,hc_kab_cirebon,tur_kab_cirebon,row_kab_garut,hp_kab_garut,hc_kab_garut,tur_kab_garut,row_kab_indramayu,hp_kab_indramayu,hc_kab_indramayu,tur_kab_indramayu,row_kab_karawang,hp_kab_karawang,hc_kab_karawang,tur_kab_karawang,row_kab_kuningan,hp_kab_kuningan,hc_kab_kuningan,tur_kab_kuningan,row_kab_majalengka,hp_kab_majalengka,hc_kab_majalengka,tur_kab_majalengka,row_kab_pangandaran,hp_kab_pangandaran,hc_kab_pangandaran,tur_kab_pangandaran,row_kab_puwakarta,hp_kab_puwakarta,hc_kab_puwakarta,tur_kab_puwakarta,row_kab_subang,hp_kab_subang,hc_kab_subang,tur_kab_subang,row_kab_sukabumi,hp_kab_sukabumi,hc_kab_sukabumi,tur_kab_sukabumi,row_kab_sumedang,hp_kab_sumedang,hc_kab_sumedang,tur_kab_sumedang,row_kab_tasikmalaya, hp_kab_tasikmalaya, hc_kab_tasikmalaya, tur_kab_tasikmalaya, row_kota_bandung, hp_kota_bandung, hc_kota_bandung, tur_kota_bandung, row_kota_banjar, hp_kota_banjar, hc_kota_banjar ,tur_kota_banjar, row_kab_cimahi, hp_kab_cimahi, hc_kab_cimahi, tur_kab_cimahi FROM map_statusfat")
        map1 = cursor.fetchall()

        cursor.execute("SELECT * FROM maptur")
        map = cursor.fetchall()

        cursor.close()
        mydb.close()

        return render_template("menu/user/mapikruser.html",map=map,map1=map1)
    except Exception as e:
        return str(e), 500




@app.route("/mapikradmin")
def mapikradmin():
    # Establish a database connection
    mydb = get_db_connection()

    if mydb is None:
        # Handle the case where the connection couldn't be established
        return "Unable to connect to the database", 500

    cursor = mydb.cursor(dictionary=True)

    try:
        cursor.execute("SELECT row_kab_bandung, hp_kab_bandung,hc_kab_bandung, tur_kab_bandung, row_kab_bandungbarat, hp_kab_bandungbarat,hc_kab_bandungbarat,tur_kab_bandungbarat,row_kab_ciamis,hp_kab_ciamis,hc_kab_ciamis,tur_kab_ciamis,row_kab_cianjur,hp_kab_cianjur,hc_kab_cianjur,tur_kab_cianjur,row_kab_cirebon,hp_kab_cirebon,hc_kab_cirebon,tur_kab_cirebon,row_kab_garut,hp_kab_garut,hc_kab_garut,tur_kab_garut,row_kab_indramayu,hp_kab_indramayu,hc_kab_indramayu,tur_kab_indramayu,row_kab_karawang,hp_kab_karawang,hc_kab_karawang,tur_kab_karawang,row_kab_kuningan,hp_kab_kuningan,hc_kab_kuningan,tur_kab_kuningan,row_kab_majalengka,hp_kab_majalengka,hc_kab_majalengka,tur_kab_majalengka,row_kab_pangandaran,hp_kab_pangandaran,hc_kab_pangandaran,tur_kab_pangandaran,row_kab_puwakarta,hp_kab_puwakarta,hc_kab_puwakarta,tur_kab_puwakarta,row_kab_subang,hp_kab_subang,hc_kab_subang,tur_kab_subang,row_kab_sukabumi,hp_kab_sukabumi,hc_kab_sukabumi,tur_kab_sukabumi,row_kab_sumedang,hp_kab_sumedang,hc_kab_sumedang,tur_kab_sumedang,row_kab_tasikmalaya, hp_kab_tasikmalaya, hc_kab_tasikmalaya, tur_kab_tasikmalaya, row_kota_bandung, hp_kota_bandung, hc_kota_bandung, tur_kota_bandung, row_kota_banjar, hp_kota_banjar, hc_kota_banjar ,tur_kota_banjar, row_kab_cimahi, hp_kab_cimahi, hc_kab_cimahi, tur_kab_cimahi FROM map_statusfat")
        map1 = cursor.fetchall()

        cursor.execute("SELECT * FROM maptur")
        map = cursor.fetchall()

        cursor.close()
        mydb.close()

        return render_template("menu/admin/mapikr.html",map=map,map1=map1)
    except Exception as e:
        return str(e), 500

    

@app.route("/datamap")
def datamap():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT label_up3_purwakarta,hp_purwakarta,hc_purwakarta,tur_purwakarta,label_up3_cirebon,hp_cirebon,hc_cirebon,tur_cirebon,label_up3_cimahi,hp_cimahi,hc_cimahi,tur_cimahi,label_up3_sumedang,hp_sumedang,hc_sumedang,tur_sumedang,label_up3_garut,hp_garut,hc_garut,tur_garut,label_up3_majalaya,hp_majalaya,hc_majalaya,tur_majalaya,label_up3_indramayu,hp_indramayu,hc_indramayu,tur_indramayu,label_up3_tasikmalaya,hp_tasikmalaya,hc_tasikmalaya,tur_tasikmalaya,label_up3_cianjur,hp_cianjur,hc_cianjur,tur_cianjur,label_up3_bandung,hp_bandung,hc_bandung,tur_bandung,label_up3_karawang,hp_karawang,hc_karawang,tur_karawang,label_up3_sukabumi,hp_sukabumi,hc_sukabumi,tur_sukabumi FROM maptur")
    datamap = cursor.fetchall()
    cursor.close()
    return render_template("tb_admin/tb_assets/ikr.html",datamap=datamap)

# New route to delete all data
@app.route("/delete_all_datamap", methods=['POST'])
def delete_all_datamap():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM maptur")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('datamap'))

@app.route("/export_tb_datamap", methods=['GET'])
def export_tb_datamap():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT label_up3_purwakarta,hp_purwakarta,hc_purwakarta,tur_purwakarta,label_up3_cirebon,hp_cirebon,hc_cirebon,tur_cirebon,label_up3_cimahi,hp_cimahi,hc_cimahi,tur_cimahi,label_up3_sumedang,hp_sumedang,hc_sumedang,tur_sumedang,label_up3_garut,hp_garut,hc_garut,tur_garut,label_up3_majalaya,hp_majalaya,hc_majalaya,tur_majalaya,label_up3_indramayu,hp_indramayu,hc_indramayu,tur_indramayu,label_up3_tasikmalaya,hp_tasikmalaya,hc_tasikmalaya,tur_tasikmalaya,label_up3_cianjur,hp_cianjur,hc_cianjur,tur_cianjur,label_up3_bandung,hp_bandung,hc_bandung,tur_bandung,label_up3_karawang,hp_karawang,hc_karawang,tur_karawang,label_up3_sukabumi,hp_sukabumi,hc_sukabumi,tur_sukabumi FROM maptur")
    dailyhcspa = cursor.fetchall()
    cursor.close()

    # Create a Pandas DataFrame from the database query
    df = pd.DataFrame(dailyhcspa)

    # Create an in-memory Excel file
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)

    # Set up response headers for Excel file download
    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=tb_datamap_export.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route("/import_tb_datamap", methods=['POST'])
def import_tb_datamap():
    file = request.files['file']
    if file:
        # Read the Excel file into a Pandas DataFrame
        df = pd.read_excel(file)

        # Ensure that the DataFrame columns match the expected columns
        expected_columns = ["label_up3_purwakarta","hp_purwakarta","hc_purwakarta","tur_purwakarta","label_up3_cirebon","hp_cirebon","hc_cirebon","tur_cirebon","label_up3_cimahi","hp_cimahi","hc_cimahi","tur_cimahi","label_up3_sumedang","hp_sumedang","hc_sumedang","tur_sumedang","label_up3_garut","hp_garut","hc_garut","tur_garut","label_up3_majalaya","hp_majalaya","hc_majalaya","tur_majalaya","label_up3_indramayu","hp_indramayu","hc_indramayu","tur_indramayu","label_up3_tasikmalaya","hp_tasikmalaya","hc_tasikmalaya","tur_tasikmalaya","label_up3_cianjur","hp_cianjur","hc_cianjur","tur_cianjur","label_up3_bandung","hp_bandung","hc_bandung","tur_bandung","label_up3_karawang","hp_karawang","hc_karawang","tur_karawang","label_up3_sukabumi","hp_sukabumi","hc_sukabumi","tur_sukabumi"]

        if not all(column in df.columns for column in expected_columns):
            return "Error: Columns in the uploaded file do not match the expected columns."

        # Handle NaN values in the DataFrame by replacing them with None
        df = df.where(pd.notna(df), None)

        # Convert int32 and int64 values to Python Decimal for better MySQL compatibility
        for col in expected_columns:
            if np.issubdtype(df[col].dtype, np.integer):
                df[col] = df[col].astype(int).apply(lambda x: Decimal(x) if x is not None else None)

        # Convert float64 values to Python Decimal for better MySQL compatibility
        for col in expected_columns:
            if df[col].dtype == np.float64:
                df[col] = df[col].apply(lambda x: Decimal(x) if not pd.isna(x) else None)

        # You can now process the DataFrame as needed, for example, insert into the database
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Assuming 'tb_dailyhcspa' is the table name
        # Using executemany to insert multiple rows at once
        columns_str = ', '.join(expected_columns)
        placeholders = ', '.join(['%s' for _ in expected_columns])
        query = f"INSERT INTO maptur ({columns_str}) VALUES ({placeholders})"
        
        try:
            # Convert DataFrame values to a list of tuples
            data_to_insert = [tuple(row) for row in df.values]

            cursor.executemany(query, data_to_insert)
            mydb.commit()
            return "File uploaded and data imported successfully."
        except Exception as e:
            return f"Error: {e}"
        finally:
            cursor.close()

    return "No file provided."



@app.route("/pdtfat")
def pdtfat():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)

    cursor.execute("SELECT row_kab_bandung,hp_kab_bandung,hc_kab_bandung,tur_kab_bandung,row_kab_bandungbarat,hp_kab_bandungbarat,hc_kab_bandungbarat,tur_kab_bandungbarat,row_kab_ciamis,hp_kab_ciamis,hc_kab_ciamis,tur_kab_ciamis,row_kab_cianjur,hp_kab_cianjur,hc_kab_cianjur,tur_kab_cianjur,row_kab_cirebon,hp_kab_cirebon,hc_kab_cirebon,tur_kab_cirebon,row_kab_garut,hp_kab_garut,hc_kab_garut,tur_kab_garut,row_kab_indramayu,hp_kab_indramayu,hc_kab_indramayu,tur_kab_indramayu,row_kab_karawang,hp_kab_karawang,hc_kab_karawang,tur_kab_karawang,row_kab_kuningan,hp_kab_kuningan,hc_kab_kuningan,tur_kab_kuningan,row_kab_majalengka,hp_kab_majalengka,hc_kab_majalengka,tur_kab_majalengka,row_kab_pangandaran,hp_kab_pangandaran,hc_kab_pangandaran,tur_kab_pangandaran,row_kab_puwakarta,hp_kab_puwakarta,hc_kab_puwakarta,tur_kab_puwakarta,row_kab_subang,hp_kab_subang,hc_kab_subang,tur_kab_subang,row_kab_sukabumi,hp_kab_sukabumi,hc_kab_sukabumi,tur_kab_sukabumi,row_kab_sumedang,hp_kab_sumedang,hc_kab_sumedang,tur_kab_sumedang,row_kab_tasikmalaya,hp_kab_tasikmalaya,hc_kab_tasikmalaya,tur_kab_tasikmalaya,row_kota_bandung,hp_kota_bandung,hc_kota_bandung,tur_kota_bandung,row_kota_banjar,hp_kota_banjar,hc_kota_banjar,tur_kota_banjar,row_kab_cimahi,hp_kab_cimahi,hc_kab_cimahi,tur_kab_cimahi FROM map_statusfat")
    fat = cursor.fetchall()
    cursor.close()
    return render_template("menu/user/fat.html",fat=fat)


# New route to delete all data
@app.route("/delete_all_pdtfat", methods=['POST'])
def delete_all_pdtfat():
    try:
        # Get the database connection
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Execute the SQL query to delete all rows from tb_olt
        cursor.execute("DELETE FROM map_statusfat")
        mydb.commit()

        # Close the cursor and the database connection
        cursor.close()
        mydb.close()

        # Provide a success message to be displayed
        flash("All data in tb_olt has been deleted successfully.", "success")

    except Exception as e:
        # Log the exception or handle it in a way that makes sense for your application
        flash("An error occurred while deleting all data. Please try again later.", "error")

    # Redirect back to the tb_olt route
    return redirect(url_for('pdtfat'))

@app.route("/export_tb_pdtfat", methods=['GET'])
def export_tb_pdtfat():
    mydb = get_db_connection()
    cursor = mydb.cursor(dictionary=True)
    cursor.execute("SELECT row_kab_bandung,hp_kab_bandung,hc_kab_bandung,tur_kab_bandung,row_kab_bandungbarat,hp_kab_bandungbarat,hc_kab_bandungbarat,tur_kab_bandungbarat,row_kab_ciamis,hp_kab_ciamis,hc_kab_ciamis,tur_kab_ciamis,row_kab_cianjur,hp_kab_cianjur,hc_kab_cianjur,tur_kab_cianjur,row_kab_cirebon,hp_kab_cirebon,hc_kab_cirebon,tur_kab_cirebon,row_kab_garut,hp_kab_garut,hc_kab_garut,tur_kab_garut,row_kab_indramayu,hp_kab_indramayu,hc_kab_indramayu,tur_kab_indramayu,row_kab_karawang,hp_kab_karawang,hc_kab_karawang,tur_kab_karawang,row_kab_kuningan,hp_kab_kuningan,hc_kab_kuningan,tur_kab_kuningan,row_kab_majalengka,hp_kab_majalengka,hc_kab_majalengka,tur_kab_majalengka,row_kab_pangandaran,hp_kab_pangandaran,hc_kab_pangandaran,tur_kab_pangandaran,row_kab_puwakarta,hp_kab_puwakarta,hc_kab_puwakarta,tur_kab_puwakarta,row_kab_subang,hp_kab_subang,hc_kab_subang,tur_kab_subang,row_kab_sukabumi,hp_kab_sukabumi,hc_kab_sukabumi,tur_kab_sukabumi,row_kab_sumedang,hp_kab_sumedang,hc_kab_sumedang,tur_kab_sumedang,row_kab_tasikmalaya,hp_kab_tasikmalaya,hc_kab_tasikmalaya,tur_kab_tasikmalaya,row_kota_bandung,hp_kota_bandung,hc_kota_bandung,tur_kota_bandung,row_kota_banjar,hp_kota_banjar,hc_kota_banjar,tur_kota_banjar,row_kab_cimahi,hp_kab_cimahi,hc_kab_cimahi,tur_kab_cimahi FROM map_statusfat")
    dailyhcspa = cursor.fetchall()
    cursor.close()

    # Create a Pandas DataFrame from the database query
    df = pd.DataFrame(dailyhcspa)

    # Create an in-memory Excel file
    excel_data = BytesIO()
    df.to_excel(excel_data, index=False)

    # Set up response headers for Excel file download
    response = make_response(excel_data.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=tb_datamap2pdtfat_export.xlsx'
    response.headers['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response


@app.route("/import_tb_pdtfat", methods=['POST'])
def import_tb_pdtfat():
    file = request.files['file']
    if file:
        # Read the Excel file into a Pandas DataFrame
        df = pd.read_excel(file)

        # Ensure that the DataFrame columns match the expected columns
        expected_columns = ["row_kab_bandung","hp_kab_bandung","hc_kab_bandung","tur_kab_bandung","row_kab_bandungbarat","hp_kab_bandungbarat","hc_kab_bandungbarat","tur_kab_bandungbarat","row_kab_ciamis","hp_kab_ciamis","hc_kab_ciamis","tur_kab_ciamis","row_kab_cianjur","hp_kab_cianjur","hc_kab_cianjur","tur_kab_cianjur","row_kab_cirebon","hp_kab_cirebon","hc_kab_cirebon","tur_kab_cirebon","row_kab_garut","hp_kab_garut","hc_kab_garut","tur_kab_garut","row_kab_indramayu","hp_kab_indramayu","hc_kab_indramayu","tur_kab_indramayu","row_kab_karawang","hp_kab_karawang","hc_kab_karawang","tur_kab_karawang","row_kab_kuningan","hp_kab_kuningan","hc_kab_kuningan","tur_kab_kuningan","row_kab_majalengka","hp_kab_majalengka","hc_kab_majalengka","tur_kab_majalengka","row_kab_pangandaran","hp_kab_pangandaran","hc_kab_pangandaran","tur_kab_pangandaran","row_kab_puwakarta","hp_kab_puwakarta","hc_kab_puwakarta","tur_kab_puwakarta","row_kab_subang","hp_kab_subang","hc_kab_subang","tur_kab_subang","row_kab_sukabumi","hp_kab_sukabumi","hc_kab_sukabumi","tur_kab_sukabumi","row_kab_sumedang","hp_kab_sumedang","hc_kab_sumedang","tur_kab_sumedang","row_kab_tasikmalaya","hp_kab_tasikmalaya","hc_kab_tasikmalaya","tur_kab_tasikmalaya","row_kota_bandung","hp_kota_bandung","hc_kota_bandung","tur_kota_bandung","row_kota_banjar","hp_kota_banjar","hc_kota_banjar","tur_kota_banjar","row_kab_cimahi","hp_kab_cimahi","hc_kab_cimahi","tur_kab_cimahi"]

        if not all(column in df.columns for column in expected_columns):
            return "Error: Columns in the uploaded file do not match the expected columns."

        # Handle NaN values in the DataFrame by replacing them with None
        df = df.where(pd.notna(df), None)

        # Convert int32 and int64 values to Python Decimal for better MySQL compatibility
        for col in expected_columns:
            if np.issubdtype(df[col].dtype, np.integer):
                df[col] = df[col].astype(int).apply(lambda x: Decimal(x) if x is not None else None)

        # Convert float64 values to Python Decimal for better MySQL compatibility
        for col in expected_columns:
            if df[col].dtype == np.float64:
                df[col] = df[col].apply(lambda x: Decimal(x) if not pd.isna(x) else None)

        # You can now process the DataFrame as needed, for example, insert into the database
        mydb = get_db_connection()
        cursor = mydb.cursor()

        # Assuming 'tb_dailyhcspa' is the table name
        # Using executemany to insert multiple rows at once
        columns_str = ', '.join(expected_columns)
        placeholders = ', '.join(['%s' for _ in expected_columns])
        query = f"INSERT INTO  map_statusfat ({columns_str}) VALUES ({placeholders})"
        
        try:
            # Convert DataFrame values to a list of tuples
            data_to_insert = [tuple(row) for row in df.values]

            cursor.executemany(query, data_to_insert)
            mydb.commit()
            return "File uploaded and data imported successfully."
        except Exception as e:
            return f"Error: {e}"
        finally:
            cursor.close()

    return "No file provided."


@app.route("/statistik")
def statistik():
    return render_template("error.html")

@app.route("/container")
def container():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT akumulasihcvsspa_bulan, akumulasihcvsspa_hc, akumulasihcvsspa_target , akumulasihcvsspa_capaian FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/dailyhcbdg")
def dailyhcbdg():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT tglhcbdg, hc_hcbdg FROM tb_dailyhcspa"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/monthlyhcbdg")
def monthlyhcbdg():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT monthlyhc_bulan, monthlyhc_bdg FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/acumhcbdg")
def acumhcbdg():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT akumulasihckp_bulan, akumulasihckp_bdg FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/monthlyagingbdg")
def monthlyagingbdg():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT agingkp_bulan, agingkp_bdg FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/dailyhccrb")
def dailyhccrb():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT tglhccrb, hc_hccrb FROM tb_dailyhcspa"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/monthlyhcrb")
def monthlyhcrb():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT monthlyhc_bulan, monthlyhc_crb FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/acumhccrb")
def acumhccrb():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT akumulasihckp_bulan, akumulasihckp_crb FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/monthlyagingcrb")
def monthlyagingcrb():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT agingkp_bulan, agingkp_crb FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/dailyhctsk")
def dailyhctsk():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT tglhctsk, hc_hctsk FROM tb_dailyhcspa"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/monthlyhctsk")
def monthlyhctsk():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT monthlyhc_bulan, monthlyhc_tsk FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/acumhctsk")
def acumhctsk():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT akumulasihckp_bulan, akumulasihckp_crb FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/monthlyagingtsk")
def monthlyagingtsk():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT agingkp_bulan, agingkp_tsk FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/monthly")
def monthly():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT agingkpbai_bulan, agingkpbai_averageaging FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/hcvsspa")
def hcvsspa():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT monthly_bulan, monthly_spa, monthly_hc  FROM tb_akv"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/oavshc")
def oavshc():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT tgl, qty, hc FROM tb_dailyhcspa"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/sale5")
def sale5():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data dengan dua digit pertama dari kolom 'rev'
    query = "SELECT montly_rev_bulan, CAST(SUBSTRING(montly_rev, 1, 2) AS UNSIGNED) AS montly_rev FROM tb_retail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/sale4")
def sale4():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT tgl, rev FROM tb_checkretail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)



@app.route("/bandwidth")
def bandwidth():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT bulan_badwhith, hc_badwhith FROM tb_retail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/revdaily")
def revdaily():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT tglbdg, revbdg FROM tb_checkretail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/revaccum")
def revaccum():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT bulanbdg, revbdg, accum_revbdg FROM tb_retail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/revdailycrb")
def revdailycrb():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT tglcrb, revcrb FROM tb_checkretail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/revaccumcrb")
def revaccumcrb():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT bulancrb, revcrb, accum_revcrb FROM tb_retail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/revdailytsk")
def revdailytsk():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT tgltsk, revtsk FROM tb_checkretail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)

@app.route("/revaccumtsk")
def revaccumtsk():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT bulantsk, revtsk, accum_revtsk FROM tb_retail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)



@app.route("/revaccum1")
def revaccum1():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT bulan, accumrev FROM tb_retail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/salesbtg")
def salesbtg():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT montly_bulan, montly_rev FROM tb_retail"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)




@app.route("/turcb1")
def turcb1():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()

    # Query untuk mengambil data cb1
    query = "SELECT kabcb1, capcb1, hpcb1, hccb1, sumofturcb1 FROM tb_akv_assets"
    cursor.execute(query)
    data = cursor.fetchall()
    return jsonify(data)


@app.route("/turcb2")
def turcb2():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kabcb2, capcb2, hpcb2, hccb2, sumofturcb2 FROM tb_akv_assets"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/turcb3")
def turcb3():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kabcb3, capcb3, hpcb3, hccb3, sumofturcb3 FROM tb_akv_assets"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/data1")
def data1():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kabbdg, hpbdg, hcbdg, turbdg FROM tb_tur"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/data2")
def data2():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kabcrb, hpcrb, hccrb, turcrb FROM tb_tur"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/data3")
def data3():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kabtsk, hptsk, hctsk, turtsk FROM tb_tur"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/tk1")
def tk1():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kabtk1, captk1, hptk1, hctk1, sumofturtk1 FROM tb_akv_assets"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/rtk2")
def rtk2():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kabtk2, captk2, hptk2, hctk2, sumofturtk2 FROM tb_akv_assets"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/rtk4")
def rtk4():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kabtk4, captk4, hptk4, hctk4, sumofturtk2 FROM tb_akv_assets"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)



@app.route("/invest")
def invest():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT kab_invest, captk_invest, hptk_invest, hctk_invest, sumofturtk_invest FROM tb_akv_assets"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/dgrmassets1")
def wecleyhomepass():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT bulan_tgl, sumofkafasitas FROM weeklyhompass"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


@app.route("/dgrmassets2")
def dgrmassets2():
    # Buat koneksi ke MySQL
    mydb = get_db_connection()
    cursor = mydb.cursor()
    # Query untuk mengambil data
    query = "SELECT bulan, akumulasi, target, pencapaian FROM acumhompass"
    cursor.execute(query)
    data = cursor.fetchall()

    # Tutup koneksi dan kirim data sebagai respons JSON
    cursor.close()
    mydb.close()
    return jsonify(data)


# -------------------diagram assets----------------





#   ---------------------endCrud-------------------------


# -----------------------Table admin-----------------------
@app.route('/no_results')
def no_results():
    return render_template("no_results.html")

# Logout
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


if __name__ == "__main__":
    app.run(debug=False)
